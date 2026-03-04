#!/usr/bin/env python3
import random
from datetime import date, timedelta
from table_generator import TableGenerator

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Erro: openpyxl não instalado. Instale com 'pip install openpyxl'")
    exit(1)

def save_multi_sheet_excel(tables, filepath):
    """Salva múltiplas tabelas em um único arquivo Excel (uma por aba)."""
    wb = Workbook()
    # Remove a aba padrão criada automaticamente
    wb.remove(wb.active)
    
    for table in tables:
        ws = wb.create_sheet(title=table.name[:31])
        
        style = table.style

        # Estilos
        header_font = Font(bold=style.header_bold, color=style.header_font_color)
        header_fill = PatternFill(start_color=style.header_bg_color,
                                   end_color=style.header_bg_color,
                                   fill_type="solid")
        alt_fill = PatternFill(start_color=style.alternate_row_color,
                                end_color=style.alternate_row_color,
                                fill_type="solid")
        thin_border = Border(
            left=Side(style='thin', color=style.border_color),
            right=Side(style='thin', color=style.border_color),
            top=Side(style='thin', color=style.border_color),
            bottom=Side(style='thin', color=style.border_color)
        )
        center_align = Alignment(horizontal='center', vertical='center')

        # Cabeçalho
        for col_idx, col_name in enumerate(table.columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border

        # Dados
        for row_idx, row_data in enumerate(table.rows, 2):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='left', vertical='center')
                # Cor alternada
                if row_idx % 2 == 0:
                    cell.fill = alt_fill

        # Auto-ajuste de largura
        if style.auto_width:
            for col_idx, col_name in enumerate(table.columns, 1):
                max_length = len(str(col_name))
                for row in table.rows:
                    if col_idx <= len(row):
                        max_length = max(max_length, len(str(row[col_idx - 1])))
                ws.column_dimensions[get_column_letter(col_idx)].width = max_length + 2

    if not filepath.endswith('.xlsx'):
        filepath += '.xlsx'
    
    wb.save(filepath)
    print(f"Arquivo '{filepath}' criado com sucesso contendo {len(tables)} abas.")

def generate_exercise_3():
    """
    Gera o arquivo 'TechStore_Desafios.xlsx' para o Exercício 3 do guia full-speed-learning.md.
    Inclui as abas 'Vendas' e 'Catálogo' pré-preenchidas.
    """
    gen = TableGenerator()

    # --- 1. Preparar Dados do Catálogo ---
    # Produto | Categoria | Fornecedor | Margem_Lucro | Preço Base (auxiliar)
    catalog_data = [
        ("Fone Bluetooth", "Eletrônicos", "FornecedorA", "35%", 150.00),
        ("Carregador USB", "Eletrônicos", "FornecedorA", "40%", 50.00),
        ("Capa de Celular", "Acessórios", "FornecedorB", "55%", 30.00),
        ("Película", "Acessórios", "FornecedorB", "60%", 20.00),
        ("Mouse Wireless", "Informática", "FornecedorC", "30%", 80.00),
        ("Teclado Mecânico", "Informática", "FornecedorC", "25%", 350.00),
    ]

    # Criar tabela Catálogo
    # Colunas: Produto, Categoria, Fornecedor, Margem_Lucro
    gen.process("crie tabela Catálogo com colunas: Produto, Categoria, Fornecedor, Margem_Lucro")
    
    for item in catalog_data:
        # item: (prod, cat, forn, margem, price)
        # Adicionamos apenas as 4 primeiras colunas na tabela
        row_str = f"linha: {item[0]}, {item[1]}, {item[2]}, {item[3]}"
        gen.process(row_str)

    # --- 2. Preparar Dados de Vendas ---
    # Colunas: ID da Venda, Data, Vendedor, Produto, Categoria, Quantidade, Valor Unitário
    
    gen.process("crie tabela Vendas com colunas: ID da Venda, Data, Vendedor, Produto, Categoria, Quantidade, Valor Unitário")

    sellers = ["Ana", "Bruno", "Carla", "Diego"]
    start_date = date(2025, 1, 1)

    # Gerar 20 linhas de vendas
    for i in range(1, 21):
        sale_id = i
        # Data aleatória em jan/2025
        sale_date = start_date + timedelta(days=random.randint(0, 30))
        sale_date_str = sale_date.strftime("%Y-%m-%d")
        
        seller = random.choice(sellers)
        
        # Escolher produto aleatório
        prod_info = random.choice(catalog_data)
        product = prod_info[0]
        category = prod_info[1]
        price = prod_info[4]
        
        quantity = random.randint(1, 5)
        
        # Montar linha
        # ID, Data, Vendedor, Produto, Categoria, Qtde, Valor
        row_str = f"linha: {sale_id}, {sale_date_str}, {seller}, {product}, {category}, {quantity}, {price}"
        gen.process(row_str)

    # --- 3. Exportar ---
    print("Gerando arquivo 'TechStore_Desafios.xlsx'...")
    # Usamos nossa função customizada para salvar ambas as tabelas no mesmo arquivo
    save_multi_sheet_excel(gen.tables, "TechStore_Desafios")

if __name__ == "__main__":
    generate_exercise_3()
