#!/usr/bin/env python3
"""
Gera o arquivo inicial para o Exercício 5 (Desafio 5) do curso Full Speed Learning.
Cria o arquivo 'exercicio_5.xlsx' com as abas 'Vendas' e 'Catálogo' preenchidas.
"""

import sys
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    print("Erro: openpyxl não instalado. Execute: pip install openpyxl")
    sys.exit(1)

def create_exercise_file():
    wb = Workbook()
    
    # --- Aba 1: Vendas ---
    ws1 = wb.active
    ws1.title = "Vendas"
    
    # Cabeçalho Vendas
    headers_vendas = ["ID_Venda", "Data", "Vendedor", "Produto", "Categoria", "Quantidade", "Valor_Unitario"]
    
    # Dados Vendas (20 linhas)
    data_vendas = [
        [1, "03/01/2025", "Ana", "Fone Bluetooth", "Eletrônicos", 3, 89.90],
        [2, "05/01/2025", "Bruno", "Película", "Acessórios", 8, 15.00],
        [3, "06/01/2025", "Carla", "Teclado Mecânico", "Informática", 1, 249.90],
        [4, "07/01/2025", "Diego", "Capa de Celular", "Acessórios", 5, 29.90],
        [5, "08/01/2025", "Ana", "Mouse Wireless", "Informática", 2, 79.90],
        [6, "09/01/2025", "Bruno", "Fone Bluetooth", "Eletrônicos", 4, 89.90],
        [7, "10/01/2025", "Carla", "Carregador USB", "Eletrônicos", 6, 45.50],
        [8, "11/01/2025", "Diego", "Teclado Mecânico", "Informática", 1, 249.90],
        [9, "13/01/2025", "Ana", "Película", "Acessórios", 10, 15.00],
        [10, "14/01/2025", "Bruno", "Mouse Wireless", "Informática", 3, 79.90],
        [11, "15/01/2025", "Carla", "Capa de Celular", "Acessórios", 7, 29.90],
        [12, "16/01/2025", "Diego", "Fone Bluetooth", "Eletrônicos", 2, 89.90],
        [13, "18/01/2025", "Ana", "Carregador USB", "Eletrônicos", 5, 45.50],
        [14, "19/01/2025", "Bruno", "Teclado Mecânico", "Informática", 2, 249.90],
        [15, "20/01/2025", "Carla", "Fone Bluetooth", "Eletrônicos", 3, 89.90],
        [16, "22/01/2025", "Diego", "Película", "Acessórios", 9, 15.00],
        [17, "24/01/2025", "Ana", "Capa de Celular", "Acessórios", 4, 29.90],
        [18, "26/01/2025", "Bruno", "Carregador USB", "Eletrônicos", 6, 45.50],
        [19, "28/01/2025", "Carla", "Mouse Wireless", "Informática", 2, 79.90],
        [20, "30/01/2025", "Diego", "Capa de Celular", "Acessórios", 3, 29.90]
    ]
    
    # Estilização básica
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    center_align = Alignment(horizontal='center', vertical='center')
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    # Escrever Vendas
    for col, value in enumerate(headers_vendas, 1):
        cell = ws1.cell(row=1, column=col, value=value)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border
        
    for row_idx, row_data in enumerate(data_vendas, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws1.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            if isinstance(value, (int, float)):
                 cell.alignment = center_align
    
    # Ajustar largura Vendas
    for col in ws1.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws1.column_dimensions[column].width = max_length + 2

    # --- Aba 2: Catálogo ---
    ws2 = wb.create_sheet(title="Catálogo")
    
    headers_catalogo = ["Produto", "Categoria", "Fornecedor", "Margem_Lucro"]
    
    data_catalogo = [
        ["Fone Bluetooth", "Eletrônicos", "FornecedorA", 0.35],
        ["Carregador USB", "Eletrônicos", "FornecedorA", 0.40],
        ["Capa de Celular", "Acessórios", "FornecedorB", 0.55],
        ["Película", "Acessórios", "FornecedorB", 0.60],
        ["Mouse Wireless", "Informática", "FornecedorC", 0.30],
        ["Teclado Mecânico", "Informática", "FornecedorC", 0.25]
    ]
    
    # Escrever Catálogo
    for col, value in enumerate(headers_catalogo, 1):
        cell = ws2.cell(row=1, column=col, value=value)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    for row_idx, row_data in enumerate(data_catalogo, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws2.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            if col_idx == 4: # Margem de Lucro (Percentage)
                cell.number_format = '0%'

    # Ajustar largura Catálogo
    for col in ws2.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws2.column_dimensions[column].width = max_length + 2

    filename = "exercicio_5.xlsx"
    wb.save(filename)
    print(f"Arquivo '{filename}' criado com sucesso!")

if __name__ == "__main__":
    create_exercise_file()
