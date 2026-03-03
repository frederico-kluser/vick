#!/usr/bin/env python3
"""
Table Generator - Cria tabelas compatíveis com Excel e Google Sheets
a partir de instruções simples em português ou inglês.

Autor: Gerado com Claude Code
Compatível com: Excel (.xlsx), Google Sheets, CSV
"""

import re
import csv
from typing import List, Dict, Any, Optional, Tuple
from dataclasses import dataclass, field
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


@dataclass
class TableStyle:
    """Configurações de estilo para a tabela."""
    header_bg_color: str = "4472C4"
    header_font_color: str = "FFFFFF"
    header_bold: bool = True
    alternate_row_color: str = "D9E2F3"
    border_color: str = "B4C6E7"
    auto_width: bool = True


@dataclass
class Table:
    """Representa uma tabela com colunas e dados."""
    name: str = "Tabela1"
    columns: List[str] = field(default_factory=list)
    rows: List[List[Any]] = field(default_factory=list)
    style: TableStyle = field(default_factory=TableStyle)

    def add_row(self, *values) -> None:
        """Adiciona uma linha à tabela."""
        row = list(values)
        # Preenche com valores vazios se necessário
        while len(row) < len(self.columns):
            row.append("")
        self.rows.append(row[:len(self.columns)])

    def add_rows(self, rows: List[List[Any]]) -> None:
        """Adiciona múltiplas linhas à tabela."""
        for row in rows:
            self.add_row(*row)


class InstructionParser:
    """Parser de instruções em linguagem natural."""

    # Padrões para português
    PATTERNS_PT = {
        'create_table': [
            r'cri(?:e|ar)\s+(?:uma\s+)?tabela\s+(?:chamada\s+)?["\']?(\w+)["\']?\s+com\s+(?:as\s+)?colunas?[:\s]+(.+)',
            r'cri(?:e|ar)\s+(?:uma\s+)?tabela\s+(?:de\s+)?(\w+)\s+com\s+(.+)',
            r'cri(?:e|ar)\s+(?:uma\s+)?tabela\s+com\s+(?:as\s+)?colunas?[:\s]+(.+)',
            r'nova\s+tabela\s+["\']?(\w+)["\']?\s*[:\s]+(.+)',
            r'tabela\s+["\']?(\w+)["\']?\s*[:\s]+(.+)',
        ],
        'add_row': [
            r'adicion(?:e|ar)\s+(?:uma\s+)?linha[:\s]+(.+)',
            r'nova\s+linha[:\s]+(.+)',
            r'inserir\s+(?:linha)?[:\s]+(.+)',
            r'linha[:\s]+(.+)',
        ],
        'add_column': [
            r'adicion(?:e|ar)\s+(?:uma\s+)?coluna[:\s]+["\']?(\w+)["\']?',
            r'nova\s+coluna[:\s]+["\']?(\w+)["\']?',
        ],
        'set_name': [
            r'(?:nome|renomear)\s+(?:para\s+)?["\']?(\w+)["\']?',
        ],
    }

    # Padrões para inglês
    PATTERNS_EN = {
        'create_table': [
            r'create\s+(?:a\s+)?table\s+(?:called\s+)?["\']?(\w+)["\']?\s+with\s+columns?[:\s]+(.+)',
            r'create\s+(?:a\s+)?table\s+with\s+columns?[:\s]+(.+)',
            r'new\s+table\s+["\']?(\w+)["\']?\s*[:\s]+(.+)',
            r'table\s+["\']?(\w+)["\']?\s*[:\s]+(.+)',
        ],
        'add_row': [
            r'add\s+(?:a\s+)?row[:\s]+(.+)',
            r'new\s+row[:\s]+(.+)',
            r'insert\s+(?:row)?[:\s]+(.+)',
            r'row[:\s]+(.+)',
        ],
        'add_column': [
            r'add\s+(?:a\s+)?column[:\s]+["\']?(\w+)["\']?',
            r'new\s+column[:\s]+["\']?(\w+)["\']?',
        ],
        'set_name': [
            r'(?:name|rename)\s+(?:to\s+)?["\']?(\w+)["\']?',
        ],
    }

    def __init__(self, language: str = "auto"):
        """
        Inicializa o parser.

        Args:
            language: "pt", "en" ou "auto" para detecção automática
        """
        self.language = language
        # Mescla os padrões PT e EN corretamente
        self.patterns = {}
        for key in set(self.PATTERNS_PT.keys()) | set(self.PATTERNS_EN.keys()):
            self.patterns[key] = (
                self.PATTERNS_PT.get(key, []) + self.PATTERNS_EN.get(key, [])
            )

    def _parse_values(self, text: str) -> List[str]:
        """Extrai valores separados por vírgula, ponto-e-vírgula ou pipe."""
        # Remove aspas e espaços extras
        text = text.strip()

        # Tenta diferentes separadores
        for sep in [';', '|', ',']:
            if sep in text:
                values = [v.strip().strip('"\'') for v in text.split(sep)]
                return [v for v in values if v]

        # Se não encontrou separador, retorna como valor único
        return [text.strip().strip('"\'')]

    def _try_convert_value(self, value: str) -> Any:
        """Tenta converter o valor para o tipo apropriado."""
        value = value.strip()

        # Tenta converter para número
        try:
            if '.' in value or ',' in value:
                # Trata vírgula como separador decimal (pt-BR)
                return float(value.replace(',', '.'))
            return int(value)
        except ValueError:
            pass

        # Tenta converter para booleano
        if value.lower() in ('true', 'verdadeiro', 'sim', 'yes', '1'):
            return True
        if value.lower() in ('false', 'falso', 'não', 'nao', 'no', '0'):
            return False

        return value

    def parse(self, instruction: str) -> Tuple[str, Dict[str, Any]]:
        """
        Analisa uma instrução e retorna o comando e seus argumentos.

        Args:
            instruction: A instrução em linguagem natural

        Returns:
            Tupla (comando, argumentos)
        """
        original = instruction.strip()
        instruction_lower = original.lower()

        # Tenta cada tipo de comando
        for command, patterns in self.patterns.items():
            for pattern in patterns:
                match = re.search(pattern, instruction_lower, re.IGNORECASE)
                if match:
                    # Usa posições do match para extrair do texto original
                    original_match = re.search(pattern, original, re.IGNORECASE)
                    groups = original_match.groups() if original_match else match.groups()

                    if command == 'create_table':
                        if len(groups) == 2:
                            name, columns_str = groups
                        else:
                            name = "Tabela1"
                            columns_str = groups[0]
                        columns = self._parse_values(columns_str)
                        return command, {'name': name, 'columns': columns}

                    elif command == 'add_row':
                        values = self._parse_values(groups[0])
                        values = [self._try_convert_value(v) for v in values]
                        return command, {'values': values}

                    elif command == 'add_column':
                        return command, {'name': groups[0]}

                    elif command == 'set_name':
                        return command, {'name': groups[0]}

        return 'unknown', {'raw': original}


class TableGenerator:
    """Gerador de tabelas a partir de instruções."""

    def __init__(self, style: Optional[TableStyle] = None):
        """
        Inicializa o gerador.

        Args:
            style: Estilo personalizado para as tabelas
        """
        self.parser = InstructionParser()
        self.style = style or TableStyle()
        self.current_table: Optional[Table] = None
        self.tables: List[Table] = []

    def process(self, instruction: str) -> str:
        """
        Processa uma instrução e retorna uma mensagem de status.

        Args:
            instruction: Instrução em linguagem natural

        Returns:
            Mensagem indicando o resultado da operação
        """
        command, args = self.parser.parse(instruction)

        if command == 'create_table':
            self.current_table = Table(
                name=args['name'],
                columns=args['columns'],
                style=self.style
            )
            self.tables.append(self.current_table)
            return f"Tabela '{args['name']}' criada com {len(args['columns'])} colunas: {', '.join(args['columns'])}"

        elif command == 'add_row':
            if not self.current_table:
                return "Erro: Nenhuma tabela ativa. Crie uma tabela primeiro."
            self.current_table.add_row(*args['values'])
            return f"Linha adicionada: {args['values']}"

        elif command == 'add_column':
            if not self.current_table:
                return "Erro: Nenhuma tabela ativa. Crie uma tabela primeiro."
            self.current_table.columns.append(args['name'])
            # Adiciona valor vazio para cada linha existente
            for row in self.current_table.rows:
                row.append("")
            return f"Coluna '{args['name']}' adicionada"

        elif command == 'set_name':
            if not self.current_table:
                return "Erro: Nenhuma tabela ativa. Crie uma tabela primeiro."
            old_name = self.current_table.name
            self.current_table.name = args['name']
            return f"Tabela renomeada de '{old_name}' para '{args['name']}'"

        else:
            return f"Instrução não reconhecida: '{instruction}'"

    def process_batch(self, instructions: List[str]) -> List[str]:
        """
        Processa várias instruções em sequência.

        Args:
            instructions: Lista de instruções

        Returns:
            Lista de mensagens de status
        """
        return [self.process(inst) for inst in instructions]

    def create_table(self, name: str, columns: List[str]) -> Table:
        """
        Cria uma tabela programaticamente.

        Args:
            name: Nome da tabela
            columns: Lista de nomes das colunas

        Returns:
            A tabela criada
        """
        self.current_table = Table(name=name, columns=columns, style=self.style)
        self.tables.append(self.current_table)
        return self.current_table

    def to_excel(self, filepath: str, table: Optional[Table] = None) -> str:
        """
        Exporta uma tabela para arquivo Excel (.xlsx).

        Args:
            filepath: Caminho do arquivo de saída
            table: Tabela a exportar (usa a atual se não especificada)

        Returns:
            Caminho do arquivo criado
        """
        if not OPENPYXL_AVAILABLE:
            raise ImportError(
                "openpyxl não está instalado. "
                "Execute: pip install openpyxl"
            )

        table = table or self.current_table
        if not table:
            raise ValueError("Nenhuma tabela para exportar")

        wb = Workbook()
        ws = wb.active
        ws.title = table.name[:31]  # Excel limita a 31 caracteres

        style = table.style

        # Estilos
        header_font = Font(bold=style.header_bold, color=style.header_font_color)
        header_fill = PatternFill(start_color=style.header_bg_color,
                                   end_color=style.header_bg_color,
                                   fill_type="solid")
        alt_fill = PatternFill(start_color=style.alternate_row_color,
                                end_color=style.alternate_row_color,
                                fill_type="solid")
        thin_border = Border(
            left=Side(style='thin', color=style.border_color),
            right=Side(style='thin', color=style.border_color),
            top=Side(style='thin', color=style.border_color),
            bottom=Side(style='thin', color=style.border_color)
        )
        center_align = Alignment(horizontal='center', vertical='center')

        # Cabeçalho
        for col_idx, col_name in enumerate(table.columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border

        # Dados
        for row_idx, row_data in enumerate(table.rows, 2):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='left', vertical='center')
                # Cor alternada
                if row_idx % 2 == 0:
                    cell.fill = alt_fill

        # Auto-ajuste de largura
        if style.auto_width:
            for col_idx, col_name in enumerate(table.columns, 1):
                max_length = len(str(col_name))
                for row in table.rows:
                    if col_idx <= len(row):
                        max_length = max(max_length, len(str(row[col_idx - 1])))
                ws.column_dimensions[get_column_letter(col_idx)].width = max_length + 2

        # Garante extensão .xlsx
        if not filepath.endswith('.xlsx'):
            filepath += '.xlsx'

        wb.save(filepath)
        return filepath

    def to_csv(self, filepath: str, table: Optional[Table] = None,
               delimiter: str = ',', encoding: str = 'utf-8-sig') -> str:
        """
        Exporta uma tabela para arquivo CSV.

        Args:
            filepath: Caminho do arquivo de saída
            table: Tabela a exportar (usa a atual se não especificada)
            delimiter: Separador de campos (padrão: vírgula)
            encoding: Codificação do arquivo (padrão: utf-8 com BOM)

        Returns:
            Caminho do arquivo criado
        """
        table = table or self.current_table
        if not table:
            raise ValueError("Nenhuma tabela para exportar")

        # Garante extensão .csv
        if not filepath.endswith('.csv'):
            filepath += '.csv'

        with open(filepath, 'w', newline='', encoding=encoding) as f:
            writer = csv.writer(f, delimiter=delimiter)
            writer.writerow(table.columns)
            writer.writerows(table.rows)

        return filepath

    def to_dict(self, table: Optional[Table] = None) -> List[Dict[str, Any]]:
        """
        Converte a tabela para lista de dicionários.

        Args:
            table: Tabela a converter (usa a atual se não especificada)

        Returns:
            Lista de dicionários representando as linhas
        """
        table = table or self.current_table
        if not table:
            raise ValueError("Nenhuma tabela para converter")

        return [
            dict(zip(table.columns, row))
            for row in table.rows
        ]

    def export_all(self, directory: str, format: str = 'xlsx') -> List[str]:
        """
        Exporta todas as tabelas para um diretório.

        Args:
            directory: Diretório de destino
            format: Formato de exportação ('xlsx' ou 'csv')

        Returns:
            Lista de caminhos dos arquivos criados
        """
        dir_path = Path(directory)
        dir_path.mkdir(parents=True, exist_ok=True)

        files = []
        for table in self.tables:
            filename = f"{table.name}.{format}"
            filepath = str(dir_path / filename)

            if format == 'xlsx':
                files.append(self.to_excel(filepath, table))
            else:
                files.append(self.to_csv(filepath, table))

        return files


def interactive_mode():
    """Modo interativo para criar tabelas via terminal."""
    print("=" * 60)
    print("  TABLE GENERATOR - Gerador de Tabelas Excel/Sheets")
    print("=" * 60)
    print("\nComandos disponíveis:")
    print("  - crie tabela <nome> com colunas: col1, col2, col3")
    print("  - adicione linha: valor1, valor2, valor3")
    print("  - adicione coluna: nome_coluna")
    print("  - exportar xlsx <arquivo>")
    print("  - exportar csv <arquivo>")
    print("  - mostrar (mostra tabela atual)")
    print("  - ajuda")
    print("  - sair")
    print("-" * 60)

    generator = TableGenerator()

    while True:
        try:
            instruction = input("\n> ").strip()

            if not instruction:
                continue

            lower_inst = instruction.lower()

            if lower_inst in ('sair', 'exit', 'quit', 'q'):
                print("Até logo!")
                break

            elif lower_inst in ('ajuda', 'help', 'h', '?'):
                print("\nExemplos de uso:")
                print("  crie tabela vendas com colunas: produto, quantidade, preco")
                print("  adicione linha: Notebook, 5, 2500.00")
                print("  adicione linha: Mouse, 20, 89.90")
                print("  exportar xlsx vendas.xlsx")
                print("  exportar csv vendas.csv")

            elif lower_inst.startswith('exportar xlsx') or lower_inst.startswith('export xlsx'):
                parts = instruction.split(maxsplit=2)
                if len(parts) >= 3:
                    filepath = generator.to_excel(parts[2])
                    print(f"Arquivo Excel criado: {filepath}")
                else:
                    print("Uso: exportar xlsx <arquivo>")

            elif lower_inst.startswith('exportar csv') or lower_inst.startswith('export csv'):
                parts = instruction.split(maxsplit=2)
                if len(parts) >= 3:
                    filepath = generator.to_csv(parts[2])
                    print(f"Arquivo CSV criado: {filepath}")
                else:
                    print("Uso: exportar csv <arquivo>")

            elif lower_inst in ('mostrar', 'show', 'ver', 'view'):
                if generator.current_table:
                    t = generator.current_table
                    print(f"\nTabela: {t.name}")
                    print("-" * 50)
                    print(" | ".join(f"{c:15}" for c in t.columns))
                    print("-" * 50)
                    for row in t.rows:
                        print(" | ".join(f"{str(v):15}" for v in row))
                else:
                    print("Nenhuma tabela criada ainda.")

            else:
                result = generator.process(instruction)
                print(result)

        except KeyboardInterrupt:
            print("\nInterrompido pelo usuário.")
            break
        except Exception as e:
            print(f"Erro: {e}")


def main():
    """Função principal com exemplos de uso."""
    import sys

    if len(sys.argv) > 1 and sys.argv[1] in ('-i', '--interactive'):
        interactive_mode()
        return

    # Exemplo de uso programático
    print("Exemplo de uso do Table Generator:")
    print("-" * 40)

    generator = TableGenerator()

    # Usando instruções em linguagem natural
    instructions = [
        "crie tabela vendas com colunas: Produto, Quantidade, Preço Unitário, Total",
        "adicione linha: Notebook Dell, 5, 3500.00, 17500.00",
        "adicione linha: Mouse Logitech, 20, 89.90, 1798.00",
        "adicione linha: Teclado Mecânico, 15, 250.00, 3750.00",
        "adicione linha: Monitor 24pol, 8, 899.00, 7192.00",
    ]

    print("\nProcessando instruções:")
    for inst in instructions:
        result = generator.process(inst)
        print(f"  {result}")

    # Exportando
    print("\nExportando tabela...")

    try:
        xlsx_path = generator.to_excel("exemplo_vendas")
        print(f"  Excel: {xlsx_path}")
    except ImportError as e:
        print(f"  Excel: {e}")

    csv_path = generator.to_csv("exemplo_vendas")
    print(f"  CSV: {csv_path}")

    # Mostrando como dicionário
    print("\nDados como dicionário:")
    for row in generator.to_dict():
        print(f"  {row}")

    print("\n" + "=" * 40)
    print("Para modo interativo, execute:")
    print("  python table_generator.py -i")


if __name__ == "__main__":
    main()
